import os
from openpyxl import load_workbook
import datetime

from django.core.management import call_command
from django.core.management.base import BaseCommand
from django.db.models.signals import post_save
from django.db.models import Sum
from django.contrib.contenttypes.models import ContentType
from django.conf import settings
from django.contrib.auth.models import User, Group

from GeneralApp import models
from GeneralApp.controllers import ProgressDisplayer

progress_bar = ''
class Command(BaseCommand):
	help = """Run this command to restore departments cecos"""

	def add_arguments(self,parser):
		parser.add_argument(
			"-n",
			help="Limit of rows per file to parse",
		)

	def execute(self, *args, **kwargs):
		self.progress_displayer = ProgressDisplayer()

		self.progress_bar = ''
		self._rows_limit = int(kwargs['n']) if kwargs['n'] else 0
		book = load_workbook(filename=os.getcwd()+settings.MEDIA_URL+'departamentos.xlsx', read_only=True)['departamentos']
		rows = book.max_row if not self._rows_limit else self._rows_limit
		models.DepartmentCECOS.objects.all().delete()
		builk_list = []
		for irow in range(2,rows):
			row = [cell.value for cell in book[irow]]
			builk_list.append(models.DepartmentCECOS(code=int(row[0]),name=row[1]))
			print(row[0])
		models.DepartmentCECOS.objects.bulk_create(builk_list)
		print(" - {} departments parsed.".format(rows))

	def handle(self, *args, **kwargs):
		self.execute()